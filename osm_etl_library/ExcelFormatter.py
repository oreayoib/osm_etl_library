from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side
from abc import ABC, abstractmethod

# Classes for formatting excel files

class ExcelFormatter(ABC):
    @abstractmethod
    def format_excel(self, file_path: str):
        pass

class ExcelFormatterPipeline(ExcelFormatter):
    def __init__(self, file_path: str):
        self.file_path = file_path

    def format_excel(self):
        try:
            wb = load_workbook(self.file_path)

            for ws in wb.worksheets:

                ws.freeze_panes = "B1"

                # Apply filter to the first row across all columns with data
                last_col_letter = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col_letter}1"
                
                # Resize columns
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2
                    
                # Define colour scale rule for "Completeness" columns
                completeness_rule_percentile = ColorScaleRule(
                    start_type='min', start_color='FF9999',   # white
                    mid_type='percentile', mid_value=50, mid_color='FFFFCC',  # yellow
                    end_type='max', end_color='99CC99'        # green
                )

                # Define colour scale rule for "frequency" columns
                
                freq_scale_rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='FFFFFFFF',       # White
                mid_type='num', mid_value=0.2, mid_color='FFFFFFCC',           # Light Yellow
                end_type='num', end_value=1, end_color= 'FF90EE90'             # Light green
                )
            
                # Get total number of rows for dynamic range
                max_row = ws.max_row
            
                # Loop through columns
                for col in ws.columns:
                    header_cell = col[0]
                    col_index = header_cell.column
                    col_letter = get_column_letter(col_index)
                
                    # --- 1. Auto-fit column width ---
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
                
                    # --- 2. Conditional formatting if header contains "Completeness" ---
                    if header_cell.value and "completeness" in str(header_cell.value).lower():
                        cell_range = f"{col_letter}2:{col_letter}{max_row}"
                        ws.conditional_formatting.add(cell_range, completeness_rule_percentile)

                # Loop through all columns in the header row
                for col in range(1, ws.max_column + 1):
                    header_value = str(ws.cell(row=1, column=col).value).strip().lower()
                
                    if "completeness" in header_value:
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'

                # Loop through all columns in the header row
                for col in range(1, ws.max_column + 1):
                    header_value = str(ws.cell(row=1, column=col).value).strip().lower()
                
                    if "freq (excl nan)" in header_value:
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'                            

                    if "values_report" in ws.title.lower():
                        # Constants
                        start_row = 2  # Assuming header in row 1
                        category_col_index = 1  # Column A
                        
                        # Border style
                        black_bottom_border = Side(style='thin', color='000000')
                        
                        # Loop through rows and detect category changes based on Column A
                        for row in range(start_row, ws.max_row):
                            current_value = ws.cell(row=row, column=category_col_index).value
                            next_value = ws.cell(row=row + 1, column=category_col_index).value
                        
                            if current_value != next_value:
                                # Apply bottom border across all columns with data in this row
                                for col in range(1, ws.max_column + 1):
                                    cell = ws.cell(row=row, column=col)
                        
                                    # Set or preserve other borders
                                    cell.border = Border(
                                        top=cell.border.top,
                                        left=cell.border.left,
                                        right=cell.border.right,
                                        bottom=black_bottom_border
                                        )  
                            
                        # Add conditional formatting for frequency column
                    
                        # Target header to search for
                        target_header = "freq (excl NaN)"

                        # Loop through columns in the header row (row 1)
                        for col in range(1, ws.max_column + 1):
                            header_value = str(ws.cell(row=1, column=col).value).strip()
                        
                            if header_value.lower() == target_header.lower():  # case-insensitive match
                                col_letter = get_column_letter(col)
                                data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                        
                                # Apply conditional formatting to the range
                                ws.conditional_formatting.add(data_range, freq_scale_rule)
                                break  # Exit loop once the column is found
            
            wb.save(f"{self.file_path.split('.xlsx')[0]}_formatted.xlsx")
    
            return f"step: format_data_report, ✅ Excel file formatted successfully."
        
        except Exception as e:
            return f"step: format_data_report, ❌ Failed to format Excel file: {e}"
                
