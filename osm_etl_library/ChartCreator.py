import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from .DataLoader import CSVLoader

# Classes for Chart Creation

def pixels_to_column_width(pixels):
    return pixels / 7.0

def pixels_to_row_height(pixels):
    return pixels * 0.75  # 1 pixel ≈ 0.75 point

class ChartBuilder:

    def __init__(self, file_path: str, folder_path: str, area: str):
        self.file_path = file_path
        self.folder_path = folder_path
        self.area = area

    def create_charts(self):

        wb = load_workbook(self.file_path)

        ws = wb[f"{self.area}_charts"]
    
        column_values = [cell.value for cell in ws['A']]
        
        attributes = column_values[1:]

        df = CSVLoader(self.folder_path, self.area).load_dataframes()

        try:
            for i in range(len(attributes)):
                attribute = attributes[i]
                attr = df[df[attribute].isnull() == False]
                attr_value = attr[attribute].value_counts().rename_axis('Category').reset_index(name='Count')

                if len(attr_value) > 10:
                    top_10 = attr_value.head(10)
                    other_sum = int(attr_value.iloc[10:]['Count'].sum())
                    other_row = pd.DataFrame([{'Category': 'Other', 'Count': other_sum}])
                    values_df = pd.concat([top_10, other_row], ignore_index=True)
                    
                else:
                    values_df = attr_value

                # Create bar plot
                plt.figure(figsize=(5, 3))
                plt.bar(values_df['Category'], values_df['Count'], color='skyblue')
                plt.title(f'Top 10 Categories with "Other" for {attribute}')
                plt.xlabel('Category')
                plt.ylabel('Count')
                plt.xticks(rotation=45)
                plt.tight_layout()
                chart_path = fr"{self.folder_path}\charts\{self.area}\bar_chart_{i}.png"
                plt.savefig(chart_path)
                plt.close()
            
                img = XLImage(fr"{self.folder_path}\charts\{self.area}\bar_chart_{i}.png")
            
                img.width = 140*3  
                img.height = 100*3
            
                img.anchor = f'B{2 + i}'  # Add space between images
            
                # Resize column C
                ws.column_dimensions['B'].width = pixels_to_column_width(img.width)
            
                # Resize row 2
                ws.row_dimensions[2 + i].height = pixels_to_row_height(img.height)
            
                ws.add_image(img)
            
            wb.save(fr"{self.folder_path}\reports\osm_category_charts.xlsx")                 
            print(f"step: add_graphs_to_plot_sheet, ✅ Excel file updated successfully for {self.area}.")
        
        except Exception as e:
            print(f"step: add_graphs_to_plot_sheet, ❌ Error updating Excel file '{self.file_path}': {e}")
        
class ChartBuilderPipeline:

    def __init__(self, folder_path, areas: list):
        self.folder_path = folder_path
        self.areas = areas
        
    def run_transforms(self) -> dict:

        file_path = fr"{self.folder_path}\reports\osm_category_charts.xlsx"

        try:
            for area in self.areas:
                ChartBuilder(file_path, self.folder_path, area).create_charts()
            return f"step: run_transforms, ✅ Charts created successfully for {len(self.areas)} areas."
        except Exception as e:
            return f"step: run_transforms, ❌ Error creating charts: {e}"
